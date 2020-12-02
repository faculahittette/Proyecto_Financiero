import requests
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import pandas_datareader as pdr
import datetime
from openpyxl import Workbook

print("Tenemos información de las siguientes empresas")
empresas = [
    "AAPL", "AIG", "AMZN", "AXP", "BA", "BAC", "CAJ", "CAT", "CL", "CMCSA", "COP",
    "CSCO", "CVS", "CVX", "DD", "DELL", "F", "GD", "GE", "GS", "GSK", "HD", 
    "HMC", "HPQ", "IBM", "JPM", "K", "KMB", "KO", "MCD", "MMM", "MSFT", "NAV",
    'NOC', 'NVS', 'PEP', 'PFE', 'PG', 'R', 'SAP', 'SNE', 'SNY', 'TM', 'TOT',
    'TXN', 'UN', 'VLO', 'WFC', 'WMT', 'XOM', 'XRX', 'MELI', 'GOOG',"TSLA"
]
print(empresas)

empresa1 = input("Ingrese una primer empresa para comparar: ")
emp1 = empresa1.upper()

#Para que le vuelva a pedir al usuario en caso de que no esté en la lista o esté mal ingresado
while emp1 not in empresas:
  print("Lo ingresado es erroneo")
  empresa1 = input("Ingrese una primer empresa para comparar: ")
  emp1 = empresa1.upper()

empresa2 = input("Ingrese una segunda empresa para comparar: ")
emp2 = empresa2.upper()

while emp2 not in empresas:
  print("Lo ingresado es erroneo")
  empresa2 = input("Ingrese una segunda empresa para comparar: ")
  emp2 = empresa2.upper()

empresa3 = input("Ingrese una tercera empresa para comparar: ")
emp3 = empresa3.upper()

while emp3 not in empresas:
  print("Lo ingresado es erroneo")
  empresa3 = input("Ingrese una tercera empresa para comparar: ")
  emp3 = empresa3.upper()

print(f"Las empresas a evaluar son: {emp1}, {emp2}, {emp3}")

#Descarga los datos desde yahoo
#Reseteamos index para poder acceder a la columna date
datae1 = pdr.get_data_yahoo(emp1, start=datetime.datetime(2018,1,10), end=datetime.datetime(2020,11,19))
datae2 = pdr.get_data_yahoo(emp2, start=datetime.datetime(2018,1,10), end=datetime.datetime(2020,11,19))
datae3 = pdr.get_data_yahoo(emp3, start=datetime.datetime(2018,1,10), end=datetime.datetime(2020,11,19))
rdatae1 = datae1.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')
rdatae2 = datae2.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')
rdatae3 = datae3.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')


crucex = []
crucey = []

###Para utilizarlo en el Excel
empx = ["Octubre", "Noviembre", "Últimos 12 meses"]
empy = []
NomEmp = []

##OCTUBRE
ymax1 = 0
ymin1 = 0
ymax2 = 0
ymin2 = 0
ymax3 = 0
ymin3 = 0

##NOVIEMBRE
ymax1n = 0
ymin1n = 0
ymax2n = 0
ymin2n = 0
ymax3n = 0
ymin3n = 0

##EN LOS ÚLTIMOS 12 MESES
ymax1a = 0
ymin1a = 0
ymax2a = 0
ymin2a = 0
ymax3a = 0
ymin3a = 0

###FILTRAMOS LOS RANGOS  DE DATOS PARA C/U
filter1 = datae1.loc['2020-10-01':'2020-11-01']
filter2 = datae2.loc['2020-10-01':'2020-11-01']
filter3 = datae3.loc['2020-10-01':'2020-11-01']

filter1n = datae1.loc['2020-09-01':'2020-10-01']
filter2n = datae2.loc['2020-09-01':'2020-10-01']
filter3n = datae3.loc['2020-09-01':'2020-10-01']

filter1a = datae1.loc['2019-11-18':'2020-11-19']
filter2a = datae2.loc['2019-11-18':'2020-11-19']
filter3a = datae3.loc['2019-11-18':'2020-11-19']

#Gráfico 1
cante1 = len(rdatae1["Date"])
x1 = []
y1 = []

for i in range(cante1):
  x1.append(rdatae1["Date"][i])
  y1.append(rdatae1["Open"][i])


ymax1 = filter1["Close"].max()
ymin1 = filter1["Close"].min()

ymax1n = filter1n["Close"].max()
ymin1n = filter1n["Close"].min()

ymax1a = filter1a["Close"].max()
ymin1a = filter1a["Close"].min()

dif1 = ymax1 - ymin1
dif1n = ymax1n - ymin1n
dif1a = ymax1a - ymin1a


#Gráfico 2
cante2 = len(rdatae2["Date"])
x2 = []
y2 = []

for i in range(cante2):
  x2.append(rdatae2["Date"][i])
  y2.append(rdatae2["Open"][i])

  if (y1[i] == y2[i]) or (y1[i] > y2[i] and y1[i-1] < y2[i-1]) or (y1[i] < y2[i] and y1[i-1] > y2[i-1]):
    crucex.append(x2[i])
    crucey.append(y2[i])

ymax2 = filter2["Close"].max()
ymin2 = filter2["Close"].min()

ymax2n = filter2n["Close"].max()
ymin2n = filter2n["Close"].min()

ymax2a = filter2a["Close"].max()
ymin2a = filter2a["Close"].min()

dif2 = ymax2 - ymin2 
dif2n = ymax2n - ymin2n
dif2a = ymax2a - ymin2a

#Gráfico 3
cante3 = len(rdatae3["Date"])
x3 = []
y3 = []

for i in range(cante3):
  x3.append(rdatae3["Date"][i])
  y3.append(rdatae3["Open"][i])

  if (y1[i] == y3[i]) or (y1[i] > y3[i] and y1[i-1] < y3[i-1]) or (y1[i] < y3[i] and y1[i-1] > y3[i-1]):
    crucex.append(x3[i])
    crucey.append(y3[i])

  if (y2[i] == y3[i]) or (y2[i] > y3[i] and y2[i-1] < y3[i-1]) or (y2[i] < y3[i] and y2[i-1] > y3[i-1]):
    crucex.append(x3[i])
    crucey.append(y3[i])

ymax3 = filter3["Close"].max()
ymin3 = filter3["Close"].min()

ymax3n = filter3n["Close"].max()
ymin3n = filter3n["Close"].min()

ymax3a = filter3a["Close"].max()
ymin3a = filter3a["Close"].min()

dif3 = ymax3 - ymin3
dif3n = ymax3n - ymin3n
dif3a = ymax3a - ymin3a

###RESULTADO OCTUBRE
###Comaparamos los 3 resultados y nos quedamos con el máximo
Oct = [dif1, dif2, dif3]
print(Oct)
ResultOct = max(Oct)

if ResultOct == dif1:
  print (f"La empresa que más crecio en Octubre es: {emp1}")
  #Lo agregamos a la lista
  empy.append(dif1)
  NomEmp.append(emp1)

elif ResultOct == dif2:
  print (f"La empresa que más crecio en Octubre es: {emp2}")
  empy.append(dif2)
  NomEmp.append(emp2)

else: 
  print (f"La empresa que más crecio en Octubre es: {emp3}")
  empy.append(dif3)
  NomEmp.append(emp3)

#RESULTADO NOVIEMBRE
Nov = [dif1n, dif2n, dif3n]
print(Nov)
ResultNov = max(Nov)

if ResultNov == dif1n:
  print (f"La empresa que más crecio en Noviembre es: {emp1}")
  empy.append(dif1n)
  NomEmp.append(emp1)

elif ResultNov == dif2n:
  print (f"La empresa que más crecio en Noviembre es: {emp2}")
  empy.append(dif2n)
  NomEmp.append(emp2)

else: 
  print (f"La empresa que más crecio en Noviembre es: {emp3}")
  empy.append(dif3n)
  NomEmp.append(emp3)

#ÚLTIMOS 12 MESES
doce = [dif1a, dif2a, dif3a]
print(doce)
ResultDoce = max(doce)

if ResultDoce == dif1a:
  print (f"La empresa que más crecio en los últimos 12 meses es: {emp1}")
  empy.append(dif1a)
  NomEmp.append(emp1)

elif ResultDoce == dif2a:
  print (f"La empresa que más crecio en los últimos 12 meses es: {emp2}")
  empy.append(dif2a)
  NomEmp.append(emp2)

else:
  print (f"La empresa que más crecio en los últimos 12 meses es: {emp3}")
  empy.append(dif3a)
  NomEmp.append(emp3)

#GUARDADO EN EXCEL DE LAS INTERSECCIONES
data = Workbook()
ruta = 'Intersec.xlsx'
hoja = data.active
hoja.title = "Intersecciones"

fila = 1 ###Fila donde empezamos
col_fecha = 1 ###Columna donde guardamos las fechas
col_dato = 2 ###Columna donde guardamos el dato asociado a cada fecha

###Guardamos las intersecciones en el Excel
for fecha, dato in zip(crucex, crucey):
    hoja.cell(column=col_fecha, row=fila, value=fecha.date())
    hoja.cell(column=col_dato, row=fila, value=dato)
    fila += 1

data.save(filename = ruta)


#GUARDADO EN EXCEL DE LAS EMPS MAYOR CRECIMIENTO
data2 = Workbook()
ruta2 = 'MayorCrecimiento.xlsx'
hoja2 = data2.active
hoja2.title = "Empresas"

fila2 = 1 ###Fila donde empezamos
col_mes = 1 ###Columna donde guardamos las fechas
col_dato2 = 2 ###Columna donde guardamos el dato asociados a cada fecha
col_emp = 3 ###Columna con el nombre de la empresa


for mes, empresa, nombre in zip(empx, empy, NomEmp):
    hoja2.cell(column=col_mes, row=fila2, value=mes)
    hoja2.cell(column=col_dato2, row=fila2, value=empresa)
    hoja2.cell(column=col_emp, row=fila2, value=nombre)
    fila2 += 1

data2.save(filename = ruta2)

#Mostramos Gráficos
plt.plot(x1, y1, label=emp1)
plt.plot(x2, y2, "red", label=emp2)
plt.plot(x3, y3, 'green', label=emp3)
plt.plot(crucex,crucey, 'k.')
plt.xticks(x1[ : :200], rotation=60)
plt.grid()
plt.legend()
plt.title("Análisis de Empresas")
plt.figure(figsize=(16, 8))
plt.show()

#GRÁFICO A ENTREGAR
#AMAZON, GOOGLE
import requests
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import pandas_datareader as pdr
import datetime
from openpyxl import Workbook


datae1 = pdr.get_data_yahoo('GOOG', start=datetime.datetime(2016,1,10), end=datetime.datetime(2020,11,19))
datae2 = pdr.get_data_yahoo('AMZN', start=datetime.datetime(2016,1,10), end=datetime.datetime(2020,11,19))
rgoogle = datae1.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')
ramazon = datae2.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')

print('A continuación ingrese las fechas entre las cuales desea ver las fluctuaciones de las acciones')
print('Utilice - ó / como separadores')
print("El período disponible es del 10/01/2016 al 19/11/2020")
fechaIn = str(input('Ingrese la fecha de inicio: '))
fechaCi = str(input('Ingrese la fecha de cierre: '))

##Delimitamos el df en base a lo que pide el usuario
##Realizamos un except por posibles errores en el ingreso de las fechas
##Al incluirlo en bucle se repite hasta que esté bien
while True:
  try:
    filter1n = datae1.loc[fechaIn:fechaCi]
    break #si no da error, se corta el while
  except ValueError:
    print("Ingreso mal las fechas, vuelva a ingresarlas")
    fechaIn = str(input('Ingrese la fecha de inicio: '))
    fechaCi = str(input('Ingrese la fecha de cierre: '))
fgoogle = filter1n.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')

filter2n = datae2.loc[fechaIn:fechaCi]
famazon = filter2n.reset_index(level = None , drop = False , inplace = False , col_level = 0 , col_fill = '')

##Empiezan con el valor del primer día
gx = [fgoogle["Date"][0]]
gy = [fgoogle["Open"][0]]
ax = [famazon["Date"][0]]
ay = [famazon["Open"][0]]

##Creamos una lista vacia para poder almacenar los cruces de los graficos
crucex = []
crucey = []

for i in range(1, len(filter1n)): #range empieza desde 1 en vez de 0
  gx.append(fgoogle["Date"][i])
  gy.append(fgoogle["Open"][i])
  ax.append(famazon["Date"][i])
  ay.append(famazon["Open"][i])

  # Condicional de cruce (son iguales o invirtieron su orden)
  if (ay[i] == gy[i]) or (ay[i] > gy[i] and ay[i-1] < gy[i-1]) or (ay[i] < gy[i] and ay[i-1] > gy[i-1]):
    crucex.append(gx[i])
    crucey.append(gy[i])

##Ingreso de intervalo para los label del EJE X
a = int(input('Elegir cada cuantos días quieres que aparezca un intervalo: '))

plt.figure(figsize=(16, 8)) #Tamaño del gráfico
plt.plot(gx,gy)
plt.plot(ax,ay)
plt.plot(crucex,crucey, 'k.') #Colocamos puntos en las interscciones
plt.xticks(gx[ : :a], rotation=60) # El rango de muestra lo determina el usuario
plt.savefig('AMZ-GOOG.png', transparent=True)

plt.show()